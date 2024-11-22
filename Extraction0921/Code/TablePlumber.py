### use plumber
import pdfplumber
import openpyxl
import os

from openpyxl import Workbook

def merge_cells(row, merge_rules):
    """
    Merge cells in a row based on merge rules.
    :param row: List of cell values in a row.
    :param merge_rules: List of tuples containing start and end indices to merge.
    :return: Modified row with merged cells.
    """
    for start, end in merge_rules:
        merged_text = " ".join(row[start:end+1])
        row[start:end+1] = [merged_text] + [""] * (end - start)
    return row

def analysis_table(pdf_path):
    # 打开PDF文件
    with pdfplumber.open(pdf_path) as pdf:
        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 遍历每页pdf
        for page in pdf.pages:
            # 提取表格信息
            table = page.extract_table(table_settings={
                'vertical_strategy': "text",
                "horizontal_strategy": "text"
            })
            print(table)

            # 格式化表格数据
            for row in table:
                print(row)
                # 合并特定的单元格内容
                merge_rules = [(0, 1)]  # 例如，将第0和第1个单元格合并
                row = merge_cells(row, merge_rules)
                sheet.append(row)

        workbook.save(filename="./output_tables/2.xlsx")

# 调用函数分析表格
analysis_table("../examplepdf/5.pdf")